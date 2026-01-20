from typing import List
from domain.kb_domain.serv.DocLoad.DocLoadImp.BaseLoader import BaseLoader


class XlsxLoader(BaseLoader):
    def _load_impl(self) -> List:
        """
        使用openpyxl读取.xlsx文件
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")

        content_list = []

        try:
            # 加载工作簿
            workbook = load_workbook(filename=self.file_path, read_only=True, data_only=True)

            # 遍历所有工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 遍历所有行
                for row in sheet.iter_rows(values_only=True):
                    # 过滤空值，拼接单元格内容
                    row_text = ' '.join(str(cell).strip() for cell in row if cell is not None and str(cell).strip())

                    if row_text:
                        content_list.append(row_text)

            workbook.close()

        except Exception as e:
            raise RuntimeError(f"读取.xlsx文件失败: {str(e)}")

        return content_list
